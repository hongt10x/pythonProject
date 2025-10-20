# -*- coding: UTF-8 -*-
'''
@Project ：python3
@File    ：5_openpyxl.py
@Date    ：2024/4/8 17:15
'''
import os

# 不支持读取.xls/.csv格式文件
from openpyxl import Workbook, load_workbook
import sys
import os
import uuid

def replace_excel(folder_path, file_name: str):
    name,suffix = file_name.split('.')
    excel_file_path = os.path.join(folder_path, file_name)
    import win32com.client
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    # excel.Visible = True
    # print(excel.Workbooks.Count)
    wb = excel.Workbooks.Open(excel_file_path)
    if 'xls' == suffix:
        suffix = f"{suffix}x"
    if 'csv' == suffix:
        suffix = "xlsx"
    new_file_name = f"{name}.{suffix}"
    if os.path.exists(new_file_name):
        new_file_name = f"{name}_{uuid.uuid4()}.{suffix}"
    # print(new_file_name)
    new_excel_file_path = os.sep.join([folder_path, new_file_name])

    wb.SaveAs(new_excel_file_path, FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    return new_excel_file_path,name


# filename = 'test_csv.csv'
# wb = load_workbook(filename,)
# fn = 'test_csv.csv'

# fn = 't1.xlsx'
# fn = 't1.xlsx'
fn = 't2.xls'
if len(sys.argv) > 1:
    fn = sys.argv[1]
filename,name = replace_excel(r'D:\PythonStudy\python3\230821-读取excel内容各种方法',fn)
print(filename,name)
# wb = load_workbook(filename, data_only=True)
wb = load_workbook(filename,)
print(wb.sheetnames)
try:
    sh1 = wb['Sheet1']
except:
    sh1 = wb[name]
print(sh1['A1'].value)
# print(sh1.rows)
# 获取所有行数据
for row in sh1.rows:
    # 获取所有行的列一列
    # print(row[0].value)
    for cell in row:
        print(cell.value)
# print(sh1.columns)
# for column in sh1.columns:
#     #  获取所有列的第一行
#     print(column[0].value)
#
# print(sh1.cell(row=1, column=1).value)
# print(wb,type(wb))
# sh1 = wb['Sheet1']
# 按行读取范围值
# for row in sh1.iter_rows(min_row=1,max_row=2,min_col=1,max_col=2):
#     for cell in row:
#         print(cell.value, end='\t')

# 按列读取范围值
# for row in sh1.iter_cols(min_row=1,max_row=2,min_col=1,max_col=2):
#     for cell in row:
#         print(cell.value)
print()
# 获取一定区域的值，按行读取
# for row in sh1['A1:C4']:
#     for cell in row:
#         print(cell.value)
