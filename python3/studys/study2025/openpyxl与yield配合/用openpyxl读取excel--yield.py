# -*- coding: utf-8 -*-
'''
@Software: PyCharm
@Project : python39
@File    : 用openpyxl读取excel--yield.py
@Time    : 2025/7/23 16:56
@Author  : Echo Wang
'''

from openpyxl import load_workbook


def read_excel(filepath):
    ...
    wb = load_workbook(filepath, read_only=True)

    sheet = wb['Sheet1']
    print(sheet)
    for row in sheet.iter_rows(values_only=True):
        yield row
    wb.close()


def read_excel_col(filepath, column_index):
    wb = load_workbook(filepath, read_only=True)
    sheet = wb['Sheet1']
    for row in sheet.iter_rows(values_only=True):
        if column_index < len(row):
            yield row[column_index]


if __name__ == '__main__':
    flag = 11
    if flag == 1:
        ...
        ret = read_excel('./test.xlsx')
        print(ret)
        for i in ret:
            print(i)
    flag = 11
    if flag == 1:
        ...
        ret = read_excel_col('./test.xlsx', 0)
        # print(ret, list(ret))
        for i in ret:
            print(i)
        print("#"*20)
        print("wht" in list(ret))
        print("wht4" in ['name', 'wht1', 'wht2', 'wht3', 'wht4', 'wht5', 'wht6', 'wht7', 'wht8'])
        print("#" * 20)
        for i in ret:
            print(i)
        print("#" * 20)

flag = 11
if flag == 1:
    ...


    def my_generator():
        yield "wht1"
        yield "wht2"
        yield "wht3"


    ret = my_generator()
    print(ret,list(ret),list(ret),ret,list(ret)) #执行一次list后就将生成器耗尽了
    print("wht1" in list(ret))  # False，因为生成器没有 "wht"
    print("wht1" in ['name', 'wht1', 'wht2', 'wht3', 'wht4'])  # True

flag = 11
if flag == 1:
    ...
    datas = read_excel_col('./test.xlsx', 0)
    # ret = any(x for x in datas if 'wht' == x)
    ret = any('wht3' == item for item in datas)
    print(list(datas),list(datas))
    print(ret)
    print(any((0,)))

flag = 11
if flag == 1:
    ...
    datas = read_excel_col('./test.xlsx', 0)
    # print(any(datas))
    print(any([-1j]))
    print(any((x for x in range(4) if x > 2)))
    from itertools import tee
    gen1,gen2 = tee(datas)
    x1 = [x for x in gen1]
    x2 = [x for x in gen2]
    print(x1, x2)
    print(list(gen1))

flag = 1
if flag == 1:
    ...
    datas = read_excel_col('./test.xlsx', 0)
    for i, content in enumerate(datas):
        print(i,content)
    print(datas,list(datas))