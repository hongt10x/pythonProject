# -*- coding: utf-8 -*-
'''
@Software: PyCharm
@Project : python39
@File    : baseExcel.py
@Time    : 2024/5/30 10:40
@Author  : Echo Wang
'''

from openpyxl import load_workbook
import sys
import os
import time
import inspect
import ipaddress
import copy
from common.basePath import config_path, data_path
from utils.handle_yml import get_yml_data
from utils.handle_excel import get_excel_data
from utils.handle_loguru import log


def show_time(func):
    '''时间装饰器函数'''

    def inner(*args, **kwargs):
        start_time = time.monotonic()
        # log.info(f"\n主体函数执行开始的时间: %.2f秒秒" % start_time)
        ret = func(*args, **kwargs)
        end_time = time.monotonic()
        # log.info(f"主体函数执行结束的时间: %.2f秒秒" % end_time)
        print(f"  -->函数{func.__name__}调用后，执行总耗时：%.2f秒\n" % (end_time - start_time))
        log.info(f"  -->函数{func.__name__}调用后，执行总耗时：%.2f秒\n\n" % (end_time - start_time))
        return ret

    return inner


class BaseExcel(object):
    def __init__(self, config_yml='configDatas.yml'):
        # 确认配置文件有效
        config_file = self.check_file_exist(os.path.join(config_path, config_yml))
        # 获取配置文件内容
        self.yml_data = get_yml_data(config_file)
        # 获取excel路径
        self.file_path = self.check_file_exist(os.path.join(data_path, self.yml_data['filename']))
        # 读取表格数据
        self.wb = get_excel_data(self.file_path)
        # 定位哪个类调用基类，获取类名
        self.update_sheet = self.__class__.__name__.split("_")[-1].lower()
        # print(self.update_sheet)
        self.config = self.yml_data[f"update_{self.update_sheet}"]

        # self.watcher = 1  # 表在多次操作时，为了保障保存后数据不相互覆盖

    def compare_status(self, update_sh, max_row=None):
        ...
        compare_type = inspect.stack()[0][3]
        sheet_info = inspect.stack()[1][3].split("_")[-1]
        # print(f"sheet_info: {self.configs[sheet_info]['sheetName']}")
        # print(f"compare_type: {compare_type}")
        sh = self.wb[self.config[sheet_info]['sheetName']]
        sh_max_row = sh.max_row + 1
        pattern = self.config[sheet_info][compare_type]['pattern']
        help_pattern = self.config[sheet_info][compare_type].get('help_pattern')
        compare_logic = self.config[sheet_info][compare_type].get('compare_logic')
        # print(f"configs: {self.configs[sheet_info][compare_type]}")
        compare_column = self.config[sheet_info]['compare_column']
        update_info = self.config[sheet_info][compare_type]['update_info']
        refer_index = self.config[sheet_info][compare_type]['refer_index']
        invalid_infos = [None, "#N/A", "无", "#", "-", "    "]  # 单元格内容无效字符匹配
        # print(f"update_info: {update_info}")
        print(">>>开始将{!r}中的状态信息更新到{!r}中".format(sh.title, update_sh.title))
        log.info(">>>开始将{!r}中的状体信息更新到{!r}中".format(sh.title, update_sh.title))
        # 获取索引
        pattern_indexs = self.get_index_list(pattern, sh)
        update_indexs = self.get_index_list(update_info, update_sh)
        help_pattern_indexs = self.get_index_list(help_pattern, sh)
        print(f"  查询{sh.title}表中列项数据:{pattern}<-->列项索引:{pattern_indexs}")
        print(f"  更新{update_sh.title}表中列项数据:{update_info}<-->列项索引:{update_indexs}")
        self.clean_col_values(update_indexs[compare_column:], update_info[compare_column:], update_sh)
        and_index_list = []
        cell_value_list = []
        for i in range(2, sh_max_row + 1):
            print("  与{!r}数据比对，当前执行第{}行...".format(self.config[sheet_info]['sheetName'], i))
            and_index_list.clear()
            v_index = None
            for column in range(compare_column):
                cell_value_list.clear()
                cell_value = self.get_cell_value(row=i, col=pattern_indexs[column], sh=sh)
                update_texts = self.get_col_values(update_indexs[column], update_info[column], update_sh)
                # print(f"cell_value{column}: {cell_value}")
                # print(f"update_texts{column}: {update_texts}")
                if compare_logic and compare_logic == 'AND':
                    ...
                    if cell_value not in invalid_infos and cell_value in update_texts:
                        v_index = update_texts.index(cell_value)
                        and_index_list.append(v_index)
                        # print(f"and_index_list{column}: {and_index_list}")
                        continue
                elif cell_value not in invalid_infos:
                    if cell_value in update_texts:
                        v_index = update_texts.index(cell_value)
                        if sheet_info == 'host':  # 只标记天工是否纳管
                            update_sh.cell(row=v_index + 2, column=update_indexs[-1], value="已纳管")
                            break
                    elif sheet_info == 'antivirus':
                        for value in cell_value.split():
                            if value not in invalid_infos and value in update_texts:
                                v_index = update_texts.index(value)
                    elif sheet_info == 'titan':
                        for new_index in help_pattern_indexs:
                            cell_value = self.get_cell_value(i, new_index, sh)
                            cell_value_list.append(cell_value.replace("(连接)", "") if cell_value else None)
                        cell_value_list = list(
                            filter(lambda ip: False if (ip in invalid_infos or not ip.startswith('10.')) else ip,
                                   cell_value_list))
                        cell_value_list.sort(key=lambda ip: ipaddress.ip_address(ip))
                        # print(f"sh_cell_text: {cell_value_list}")
                        if cell_value_list[column] not in invalid_infos and cell_value_list[column] in update_texts:
                            v_index = update_texts.index(cell_value_list[column])

                    # print(f"v_index:{v_index}")
                    if v_index is not None:
                        for k, idx in enumerate(update_indexs[compare_column:]):
                            cell_value = self.get_cell_value(row=i, col=pattern_indexs[compare_column + k], sh=sh)
                            update_sh.cell(row=v_index + 2, column=idx, value=cell_value)
                        break
            # print(f"and_index_list{i}: {and_index_list}")
            if and_index_list and compare_column == and_index_list.count(and_index_list[0]):
                for k, idx in enumerate(update_indexs[compare_column:]):
                    cell_value = self.get_cell_value(row=i, col=pattern_indexs[compare_column + k], sh=sh)
                    # print(f"cell_value{idx}:{cell_value}")
                    update_sh.cell(row=and_index_list[0] + 2, column=idx, value=cell_value)

        print("<<<表{!r}状态信息更新完毕".format(update_sh.title))
        log.info("<<<表{!r}状态信息更新完毕".format(update_sh.title))

    def compare_status_optimized0614(self, update_sh, max_row):
        ...
        compare_type = inspect.stack()[0][3]
        sheet_info = inspect.stack()[1][3].split("_")[-1]
        # print(f"sheet_info: {self.configs[sheet_info]['sheetName']}")
        # print(f"compare_type: {compare_type}")
        sh = self.wb[self.config[sheet_info]['sheetName']]
        sh_max_row = sh.max_row + 1
        pattern = self.config[sheet_info][compare_type]['pattern']
        help_pattern = self.config[sheet_info][compare_type].get('help_pattern')
        compare_logic = self.config[sheet_info][compare_type].get('compare_logic')
        # print(f"configs: {self.configs[sheet_info][compare_type]}")
        compare_column = self.config[sheet_info]['compare_column']
        update_info = self.config[sheet_info][compare_type]['update_info']
        refer_index = self.config[sheet_info][compare_type]['refer_index']
        invalid_infos = [None, "#N/A", "无", "#", "-", "    "]  # 单元格内容无效字符匹配
        # print(f"update_info: {update_info}")
        print(">>>开始将{!r}中的数据更新到{!r}中".format(sh.title, update_sh.title))
        log.info(">>>开始将{!r}中的数据更新到{!r}中".format(sh.title, update_sh.title))
        # 获取索引
        pattern_indexs = self.get_index_list(pattern, sh)
        update_indexs = self.get_index_list(update_info, update_sh)
        help_pattern_indexs = self.get_index_list(help_pattern, sh)
        print(f"  查询{sh.title}表中列项数据:{pattern}<-->列项索引:{pattern_indexs}")
        print(f"  更新{update_sh.title}表中列项数据:{update_info}<-->列项索引:{update_indexs}")
        self.clean_col_values(update_indexs[compare_column:], update_info[compare_column:], update_sh)
        and_index_list = []
        cell_value_list = []
        for i in range(2, sh_max_row + 1):
            print("  与{!r}数据比对，当前执行第{}行...".format(self.config[sheet_info]['sheetName'], i))
            and_index_list.clear()
            v_index = None
            for column in range(compare_column):
                cell_value_list.clear()
                cell_value = self.get_cell_value(row=i, col=pattern_indexs[column], sh=sh)
                update_texts = self.get_col_values(update_indexs[column], update_info[column], update_sh)
                # print(f"cell_value{column}: {cell_value}")
                # print(f"update_texts{column}: {update_texts}")
                if compare_logic and compare_logic == 'AND':
                    ...
                    if cell_value not in invalid_infos and cell_value in update_texts:
                        v_index = update_texts.index(cell_value)
                        and_index_list.append(v_index)
                        # print(f"and_index_list{column}: {and_index_list}")
                        continue
                elif cell_value not in invalid_infos:
                    if cell_value in update_texts:
                        v_index = update_texts.index(cell_value)
                        if sheet_info == 'host':  # 只标记天工是否纳管
                            update_sh.cell(row=v_index + 2, column=update_indexs[-1], value="已纳管")
                            break
                    elif sheet_info == 'antivirus':
                        for value in cell_value.split():
                            if value not in invalid_infos and value in update_texts:
                                v_index = update_texts.index(value)
                    elif sheet_info == 'titan':
                        for new_index in help_pattern_indexs:
                            cell_value = self.get_cell_value(i, new_index, sh)
                            cell_value_list.append(cell_value.replace("(连接)", "") if cell_value else None)
                        cell_value_list = list(
                            filter(lambda ip: False if (ip in invalid_infos or not ip.startswith('10.')) else ip,
                                   cell_value_list))
                        cell_value_list.sort(key=lambda ip: ipaddress.ip_address(ip))
                        # print(f"sh_cell_text: {cell_value_list}")
                        if cell_value_list[column] not in invalid_infos and cell_value_list[column] in update_texts:
                            v_index = update_texts.index(cell_value_list[column])

                    # print(f"v_index:{v_index}")
                    if v_index is not None:
                        for k, idx in enumerate(update_indexs[compare_column:]):
                            cell_value = self.get_cell_value(row=i, col=pattern_indexs[compare_column + k], sh=sh)
                            update_sh.cell(row=v_index + 2, column=idx, value=cell_value)
                        break
            # print(f"and_index_list{i}: {and_index_list}")
            if and_index_list and compare_column == and_index_list.count(and_index_list[0]):
                for k, idx in enumerate(update_indexs[compare_column:]):
                    cell_value = self.get_cell_value(row=i, col=pattern_indexs[compare_column + k], sh=sh)
                    # print(f"cell_value{idx}:{cell_value}")
                    update_sh.cell(row=and_index_list[0] + 2, column=idx, value=cell_value)

        print("<<<表{!r}状态更新完毕".format(update_sh.title))
        log.info("<<<表{!r}状态更新完毕".format(update_sh.title))

    def compare_status_bak0613(self, update_sh, max_row):
        ...
        compare_type = inspect.stack()[0][3]
        sheet_info = inspect.stack()[1][3].split("_")[-1]
        # print(f"sheet_info: {self.configs[sheet_info]['sheetName']}")
        # print(f"compare_type: {compare_type}")
        sh = self.wb[self.config[sheet_info]['sheetName']]
        sh_max_row = sh.max_row + 1
        pattern = self.config[sheet_info][compare_type]['pattern']
        help_pattern = self.config[sheet_info][compare_type].get('help_pattern')
        compare_logic = self.config[sheet_info][compare_type].get('compare_logic')
        # print(f"configs: {self.configs[sheet_info][compare_type]}")
        compare_column = self.config[sheet_info]['compare_column']
        update_info = self.config[sheet_info][compare_type]['update_info']
        refer_index = self.config[sheet_info][compare_type]['refer_index']
        invalid_infos = [None, "#N/A", "无", "#", "-", "    "]  # 单元格内容无效字符匹配
        # print(f"update_info: {update_info}")
        print(">>>开始将{!r}中的数据更新到{!r}中".format(sh.title, update_sh.title))
        log.info(">>>开始将{!r}中的数据更新到{!r}中".format(sh.title, update_sh.title))
        # 获取索引
        pattern_indexs = self.get_index_list(pattern, sh)

        update_indexs = self.get_index_list(update_info, update_sh)
        print(f"  查询{sh.title}表中列项数据:{pattern}<-->列项索引:{pattern_indexs}")
        print(f"  更新{update_sh.title}表中列项数据:{update_info}<-->列项索引:{update_indexs}")
        self.clean_col_values(update_indexs[compare_column:], update_info[compare_column:], update_sh)

        for i in range(2, max_row + 1):
            print("  与{!r}数据比对，当前执行第{}行...".format(self.config[sheet_info]['sheetName'], i))
            and_index_list = []
            exist_flag = True
            v_index = None
            for column in range(compare_column):
                cell_value = self.get_cell_value(row=i, col=update_indexs[column], sh=update_sh)
                sh_texts = self.get_col_values(pattern_indexs[column], pattern[column], sh)
                # print(f"cell_value{column}: {cell_value}")
                # print(f"sh_texts{column}: {sh_texts}")
                if compare_logic and compare_logic == 'AND':
                    ...
                    if cell_value not in invalid_infos and cell_value in sh_texts:
                        v_index = sh_texts.index(cell_value)
                        and_index_list.append(v_index)
                        continue
                else:
                    ...
                    if cell_value not in invalid_infos and cell_value in sh_texts:
                        v_index = sh_texts.index(cell_value)
                        if sheet_info == 'host':  # 只标记天工是否纳管
                            exist_flag = False
                            update_sh.cell(row=i, column=update_indexs[-1], value="已纳管")
                            break

                    elif sheet_info == 'antivirus' and cell_value not in invalid_infos:
                        for num, values in enumerate(sh_texts):
                            if values not in invalid_infos and cell_value in values.split():
                                v_index = num
                    # print(f"v_index:{v_index}")
                    if v_index is not None:
                        for k, idx in enumerate(update_indexs[compare_column:]):
                            cell_value = \
                                self.get_col_values(pattern_indexs[compare_column + k], pattern[compare_column + k],
                                                    sh)[v_index]
                            # print(f"cell_value{idx}:{cell_value}")
                            update_sh.cell(row=i, column=idx, value=cell_value)
                        break

            if exist_flag and sheet_info == 'host':  # 只标记天工是否纳管
                update_sh.cell(row=i, column=update_indexs[-1], value="未纳管")
            if and_index_list and compare_column == and_index_list.count(and_index_list[0]):
                for k, idx in enumerate(update_indexs[compare_column:]):
                    cell_value = \
                        self.get_col_values(pattern_indexs[compare_column + k], pattern[compare_column + k],
                                            sh)[and_index_list[0]]
                    # print(f"cell_value{idx}:{cell_value}")
                    update_sh.cell(row=i, column=idx, value=cell_value)

        print("<<<表{!r}状态更新完毕".format(update_sh.title))
        log.info("<<<表{!r}状态更新完毕".format(update_sh.title))

    def compare_datas_other(self, update_sh):
        compare_type = inspect.stack()[0][3]
        # print(inspect.stack()[1][3])
        sheet_info = inspect.stack()[1][3].split("_")[-1]
        # print(f"sheet_info:{sheet_info}")
        sh = self.wb[self.config[sheet_info]['sheetName']]
        refer_index = self.config[sheet_info][compare_type].get('refer_index')
        pattern = self.config[sheet_info][compare_type]['pattern']
        update_info = self.config[sheet_info][compare_type]['update_info']
        print(">>>开始检查{!r}在{!r}中缺失的数据".format(sh.title, update_sh.title))
        log.info(">>>开始检查{!r}在{!r}中缺失的数据".format(sh.title, update_sh.title))
        pattern_indexs = self.get_index_list(pattern, sh)  # 获取关键列索引
        update_indexs = self.get_index_list(update_info, update_sh)  # 获取关键列索引
        print(f"  查询{sh.title}表中列项数据:{pattern}<-->列项索引:{pattern_indexs}")
        print(f"  更新{update_sh.title}表中列项数据:{update_info}<-->列项索引:{update_indexs}")
        invalid_infos = [None, "#N/A", "无", "#", "-", "    "]  # 单元格内容无效字符匹配
        missed_ips = {}  # 缺失ip列表
        for num in range(len(pattern_indexs)):
            flag = True
            # self.get_col_values函数的索引从0开始，需要往前倒一位
            sh_texts = self.get_col_values(pattern_indexs[num], pattern[num], sh)
            update_texts = self.get_col_values(update_indexs[num], update_info[num], update_sh)
            # print(f"sh_texts:{sh_texts}")
            # print(f"update_texts:{update_texts}")

            for index, one in enumerate(sh_texts):
                if sheet_info == 'bastion':
                    ...

                if sheet_info == 'antivirus':
                    ...
                # 处理不存在或未匹配到的数据单元
                if one in invalid_infos or one not in update_texts:
                    missed_ips.update({index + 2: one})
                    # 如果首列未匹配到，则依次往后轮询，确认其他列是否都有相同元素，有则结束循环，无则循环到最后一个元素为止
                    for col_num in range(1, len(pattern_indexs)):
                        value = self.get_cell_value(row=index + 2, col=pattern_indexs[col_num] + 1, sh=sh)
                        # print(f"value:{value}")
                        update_texts = self.get_col_values(update_indexs[col_num], pattern[col_num], update_sh)

                        if value in invalid_infos or value not in update_texts:
                            flag = False
                        else:
                            # 找到数据后，删除missed_ips列表中的数据
                            del missed_ips[index + 2]
                            flag = True
                        if flag:
                            break

                    flag = True  # 内循环结束后，将flag置为True，避免后边再次对列项进行迭代循环
            if flag:
                break  # 标志位为True就结束循环
        print(f"<<<检查完成，返回缺失的数据的行号和ip信息:\n  {missed_ips} \n")
        log.info(f"<<<检查完成，返回缺失的数据的行号和ip信息:\n  {missed_ips}")
        # return missed_ips

        print(">>>开始将缺失数据，更新入{!r}的列{}中".format(update_sh.title, update_info))
        update_index = self.get_index_list(update_info, update_sh)
        # print(f"update_index:{update_index}")
        self.clean_col_values(update_index[refer_index:], update_info[refer_index:], update_sh)

        j = 1
        for num, ip in missed_ips.items():
            j += 1
            # print(self.watcher)
            update_sh.cell(row=j, column=update_index[refer_index:][0], value=ip)
            update_sh.cell(row=j, column=update_index[refer_index:][1], value=num)

        print("<<<表{!r}数据更新完毕".format(update_sh.title))
        log.info("<<<表{!r}数据更新完毕".format(update_sh.title))

    def compare_datas(self, update_sh, max_row):
        # 获取被调用函数名
        # print(inspect.stack()[1][3])
        # 获取当前函数名
        compare_type = inspect.stack()[0][3]
        sheet_info = inspect.stack()[1][3].split("_")[-1]
        sh = self.wb[self.config[sheet_info]['sheetName']]
        sh_max_row = sh.max_row
        # print(f"最大行：{sh_max_row}")
        compare_column = self.config[sheet_info]['compare_column']
        pattern = self.config[sheet_info][compare_type]['pattern']
        help_pattern = self.config[sheet_info][compare_type].get('help_pattern')
        update_info = self.config[sheet_info][compare_type]['update_info']
        refer_index = self.config[sheet_info][compare_type].get('refer_index')
        print(">>>开始准备将{!r}中独有的数据更新到{!r}中".format(sh.title, update_sh.title))
        log.info(">>>开始准备将{!r}中独有的数据更新到{!r}中".format(sh.title, update_sh.title))
        # 获取索引
        pattern_indexs = self.get_index_list(pattern, sh)
        update_indexs = self.get_index_list(update_info, update_sh)
        help_pattern_indexs = self.get_index_list(help_pattern, sh)
        # print(f"help_pattern_indexs: {help_pattern_indexs}")
        print(f"  查询{sh.title}表中列项数据:{pattern}<-->列项索引:{pattern_indexs}")
        print(f"  更新{update_sh.title}表中列项数据:{update_info}<-->列项索引:{update_indexs}")
        invalid_infos = [None, "#N/A", "无", "#", "-", "    "]  # 单元格内容无效字符匹配
        # os_types = ['Linux', 'Windows', 'AIX', 'HP UX']
        # 将防病毒库中独有的数据更新到OS表中
        record_line = 1
        new_update_texts = []
        for row in range(2, sh_max_row + 1):
            print("  与{!r}数据比对，当前执行第{}行...".format(self.config[sheet_info]['sheetName'], row))
            match_times = 0
            for column in range(compare_column):
                ...
                flag = False
                # 获取动态表比对列数据，并转为列表统一处理
                try:
                    sh_cell_text = self.get_cell_value(row, pattern_indexs[column], sh).split()
                except:
                    continue
                # 获取OS表比对列数据
                update_texts = self.get_col_values(update_indexs[column], update_info[column], update_sh)
                # print(f"sh_cell_text{column}: {sh_cell_text}")
                # print(f"update_texts{column}: {update_texts}")
                if sheet_info == 'titan' and column == 0:  # 对青藤云的多列数据重组，获取业务IP和带内IP
                    sh_cell_text.clear()
                    for new_index in help_pattern_indexs:
                        cell_text = self.get_cell_value(row, new_index, sh)
                        sh_cell_text.append(cell_text.replace("(连接)", "") if cell_text else None)

                    sh_cell_text = list(
                        filter(lambda ip: False if (ip in invalid_infos or not ip.startswith('10.')) else ip,
                               sh_cell_text))
                    sh_cell_text.sort(key=lambda ip: ipaddress.ip_address(ip))
                    # print(f"sh_cell_text: {sh_cell_text}")
                    new_update_texts = copy.deepcopy(sh_cell_text)
                    # print(f"new_update_texts{column}: {new_update_texts}")
                for cell_text in sh_cell_text:
                    if cell_text not in invalid_infos and cell_text in update_texts:
                        flag = True
                        break
                if flag:
                    break
                else:
                    match_times += 1  # 记录查询列数
            # print(f"match_times: {match_times}")
            if match_times == compare_column:
                # 未匹配到任何列数据
                for i, update_index in enumerate(update_indexs):
                    if update_index == update_indexs[-1]:
                        # 最后一列更新数据注明数据来源，行往后追加
                        update_sh.cell(row=max_row + record_line, column=update_index, value=sh.title)
                        break
                    if pattern_indexs.count(pattern_indexs[i]) > 1:  # 剔除重复比对数据不匹配情况
                        if i > pattern_indexs.index(pattern_indexs[i]):
                            continue
                    # 从第2行开始取值
                    update_value = self.get_cell_value(row=row, col=pattern_indexs[i], sh=sh)
                    # print(f"update_value00: {update_value}")
                    if sheet_info == 'titan' and new_update_texts and i in range(refer_index):
                        update_value = new_update_texts[i]
                    if sheet_info == 'antivirus' and i == refer_index:
                        sh_cell_text = list(
                            filter(lambda ip: False if (ip in invalid_infos or not ip.startswith('10.')) else ip,
                                   update_value.split()))
                        sh_cell_text.sort(key=lambda ip: ipaddress.ip_address(ip))
                        update_value = sh_cell_text[0]
                        try:
                            update_sh.cell(row=max_row + record_line, column=update_indexs[i + 1],
                                           value=sh_cell_text[1])
                        except:
                            pass
                    # print(f"update_value1: {update_value}")
                    if sheet_info == 'host' and update_value:  # 对主机表的操作系统数据拼接
                        if i == refer_index:
                            update_value = f"{self.get_cell_value(row=row, col=pattern_indexs[i - 1], sh=sh)} {update_value}"
                        # if i == 4:  # 对主机表的操作系统类型判断处理
                        #     for os_type in os_types:
                        #         if os_type.lower() in update_value.lower():
                        #             update_value = os_type
                    if sheet_info == 'automation' and i == refer_index and update_value:  # 对自动化表的主机名数据拆分
                        update_value = update_value.split("_")
                        if len(update_value) > 1:
                            update_sh.cell(row=max_row + record_line, column=update_indexs[i - 1],
                                           value=update_value[1].strip())
                        update_value = update_value[0].strip()

                    # print(f"new_update_value: {update_value}")
                    if type(update_value) is str:
                        update_value = update_value.strip()
                    update_sh.cell(row=max_row + record_line, column=update_indexs[i], value=update_value)

                record_line += 1
        print(">>>已完成将{!r}中独有的数据更新到{!r}中".format(sh.title, update_sh.title))
        log.info(">>>已完成将{!r}中独有的数据更新到{!r}中".format(sh.title, update_sh.title))

    def compare_datas_optimized0614(self, update_sh, max_row):
        # 获取被调用函数名
        # print(inspect.stack()[1][3])
        # 获取当前函数名
        compare_type = inspect.stack()[0][3]
        sheet_info = inspect.stack()[1][3].split("_")[-1]
        sh = self.wb[self.config[sheet_info]['sheetName']]
        sh_max_row = sh.max_row
        # print(f"最大行：{sh_max_row}")
        compare_column = self.config[sheet_info]['compare_column']
        pattern = self.config[sheet_info][compare_type]['pattern']
        help_pattern = self.config[sheet_info][compare_type].get('help_pattern')
        update_info = self.config[sheet_info][compare_type]['update_info']
        refer_index = self.config[sheet_info][compare_type].get('refer_index')
        print(">>>开始将{!r}中的数据更新到{!r}中".format(sh.title, update_sh.title))
        log.info(">>>开始将{!r}中的数据更新到{!r}中".format(sh.title, update_sh.title))
        # 获取索引
        pattern_indexs = self.get_index_list(pattern, sh)
        update_indexs = self.get_index_list(update_info, update_sh)
        help_pattern_indexs = self.get_index_list(help_pattern, sh)
        # print(f"help_pattern_indexs: {help_pattern_indexs}")
        print(f"  查询{sh.title}表中列项数据:{pattern}<-->列项索引:{pattern_indexs}")
        print(f"  更新{update_sh.title}表中列项数据:{update_info}<-->列项索引:{update_indexs}")
        invalid_infos = [None, "#N/A", "无", "#", "-", "    "]  # 单元格内容无效字符匹配
        # os_types = ['Linux', 'Windows', 'AIX', 'HP UX']
        # 将防病毒库中独有的数据更新到OS表中
        record_line = 1
        new_update_texts = []
        for row in range(2, sh_max_row + 1):
            print("  与{!r}数据比对，当前执行第{}行...".format(self.config[sheet_info]['sheetName'], row))
            match_times = 0
            for column in range(compare_column):
                ...
                flag = False
                # 获取动态表比对列数据，并转为列表统一处理
                try:
                    sh_cell_text = self.get_cell_value(row, pattern_indexs[column], sh).split()
                except:
                    continue
                # 获取OS表比对列数据
                update_texts = self.get_col_values(update_indexs[column], update_info[column], update_sh)
                # print(f"sh_cell_text{column}: {sh_cell_text}")
                # print(f"update_texts{column}: {update_texts}")
                if sheet_info == 'titan' and column == 0:  # 对青藤云的多列数据重组，获取业务IP和带内IP
                    sh_cell_text.clear()
                    for new_index in help_pattern_indexs:
                        cell_text = self.get_cell_value(row, new_index, sh)
                        sh_cell_text.append(cell_text.replace("(连接)", "") if cell_text else None)

                    sh_cell_text = list(
                        filter(lambda ip: False if (ip in invalid_infos or not ip.startswith('10.')) else ip,
                               sh_cell_text))
                    sh_cell_text.sort(key=lambda ip: ipaddress.ip_address(ip))
                    # print(f"sh_cell_text: {sh_cell_text}")
                    new_update_texts = copy.deepcopy(sh_cell_text)
                    # print(f"new_update_texts{column}: {new_update_texts}")
                for cell_text in sh_cell_text:
                    if cell_text not in invalid_infos and cell_text in update_texts:
                        flag = True
                        break
                if flag:
                    break
                else:
                    match_times += 1  # 记录查询列数
            # print(f"match_times: {match_times}")
            if match_times == compare_column:
                # 未匹配到任何列数据
                for i, update_index in enumerate(update_indexs):
                    if update_index == update_indexs[-1]:
                        # 最后一列更新数据注明数据来源，行往后追加
                        update_sh.cell(row=max_row + record_line, column=update_index, value=sh.title)
                        break
                    if pattern_indexs.count(pattern_indexs[i]) > 1:  # 剔除重复比对数据不匹配情况
                        if i > pattern_indexs.index(pattern_indexs[i]):
                            continue
                    # 从第2行开始取值
                    update_value = self.get_cell_value(row=row, col=pattern_indexs[i], sh=sh)
                    # print(f"update_value00: {update_value}")
                    if sheet_info == 'titan' and new_update_texts and i in range(refer_index):
                        update_value = new_update_texts[i]
                    if sheet_info == 'antivirus' and i == refer_index:
                        sh_cell_text = list(
                            filter(lambda ip: False if (ip in invalid_infos or not ip.startswith('10.')) else ip,
                                   update_value.split()))
                        sh_cell_text.sort(key=lambda ip: ipaddress.ip_address(ip))
                        update_value = sh_cell_text[0]
                        try:
                            update_sh.cell(row=max_row + record_line, column=update_indexs[i + 1],
                                           value=sh_cell_text[1])
                        except:
                            pass
                    # print(f"update_value1: {update_value}")
                    if sheet_info == 'host' and update_value:  # 对主机表的操作系统数据拼接
                        if i == refer_index:
                            update_value = f"{self.get_cell_value(row=row, col=pattern_indexs[i - 1], sh=sh)} {update_value}"
                        # if i == 4:  # 对主机表的操作系统类型判断处理
                        #     for os_type in os_types:
                        #         if os_type.lower() in update_value.lower():
                        #             update_value = os_type
                    if sheet_info == 'automation' and i == refer_index and update_value:  # 对自动化表的主机名数据拆分
                        update_value = update_value.split("_")
                        if len(update_value) > 1:
                            update_sh.cell(row=max_row + record_line, column=update_indexs[i - 1],
                                           value=update_value[1].strip())
                        update_value = update_value[0].strip()

                    # print(f"new_update_value: {update_value}")
                    if type(update_value) is str:
                        update_value = update_value.strip()
                    update_sh.cell(row=max_row + record_line, column=update_indexs[i], value=update_value)

                record_line += 1

        print("<<<表{!r}更新完毕".format(update_sh.title))
        log.info("<<<表{!r}更新完毕".format(update_sh.title))

    def compare_datas_bak0613(self, update_sh, max_row):
        # 获取被调用函数名
        # print(inspect.stack()[1][3])
        # 获取当前函数名
        compare_type = inspect.stack()[0][3]
        sheet_info = inspect.stack()[1][3].split("_")[-1]
        sh = self.wb[self.config[sheet_info]['sheetName']]
        sh_max_row = sh.max_row
        # print(f"最大行：{sh_max_row}")
        compare_column = self.config[sheet_info]['compare_column']
        pattern = self.config[sheet_info][compare_type]['pattern']
        help_pattern = self.config[sheet_info][compare_type].get('help_pattern')
        update_info = self.config[sheet_info][compare_type]['update_info']
        refer_index = self.config[sheet_info][compare_type].get('refer_index')
        print(">>>开始将{!r}中的数据更新到{!r}中".format(sh.title, update_sh.title))
        log.info(">>>开始将{!r}中的数据更新到{!r}中".format(sh.title, update_sh.title))
        # 获取索引
        pattern_indexs = self.get_index_list(pattern, sh)
        update_indexs = self.get_index_list(update_info, update_sh)
        help_pattern_indexs = self.get_index_list(help_pattern, sh)
        # print(f"help_pattern_indexs: {help_pattern_indexs}")
        print(f"  查询{sh.title}表中列项数据:{pattern}<-->列项索引:{pattern_indexs}")
        print(f"  更新{update_sh.title}表中列项数据:{update_info}<-->列项索引:{update_indexs}")
        invalid_infos = [None, "#N/A", "无", "#", "-", "    "]  # 单元格内容无效字符匹配
        os_types = ['Linux', 'Windows']
        # 将防病毒库中独有的数据更新到OS表中
        record_line = 1
        for row in range(2, sh_max_row + 1):
            print("  与{!r}数据比对，当前执行第{}行...".format(self.config[sheet_info]['sheetName'], row))
            match_times = 0
            new_update_texts = []
            for column in range(compare_column):
                ...
                flag = False
                # 获取动态表比对列数据，并转为列表统一处理
                try:
                    sh_cell_text = self.get_cell_value(row, pattern_indexs[column], sh).split()
                except:
                    continue
                # 获取OS表比对列数据
                update_texts = self.get_col_values(update_indexs[column], update_info[column], update_sh)
                # print(f"sh_cell_text{column}: {sh_cell_text}")
                # print(f"update_texts{column}: {update_texts}")
                if sheet_info == 'titan' and column == 0:  # 对青藤云的多列数据重组，获取业务IP和带内IP
                    sh_cell_text.clear()
                    for new_index in help_pattern_indexs:
                        cell_text = self.get_cell_value(row, new_index, sh)
                        sh_cell_text.append(cell_text.replace("(连接)", "") if cell_text else None)

                    sh_cell_text = list(
                        filter(lambda ip: False if (ip in invalid_infos or not ip.startswith('10.')) else ip,
                               sh_cell_text))
                    sh_cell_text.sort(key=lambda ip: ipaddress.ip_address(ip))
                    # print(f"sh_cell_text: {sh_cell_text}")
                    new_update_texts = copy.deepcopy(sh_cell_text)
                print(f"new_update_texts{column}: {new_update_texts}")
                for cell_text in sh_cell_text:
                    if cell_text not in invalid_infos and cell_text in update_texts:
                        # match_times -= 1
                        flag = False
                        break
                    else:
                        flag = True
                if flag:
                    match_times += 1  # 记录查询列数

            if match_times == compare_column:
                # 未匹配到任何列数据
                for i, update_index in enumerate(update_indexs):
                    if update_index == update_indexs[-1]:
                        # 最后一列更新数据注明数据来源，行往后追加
                        update_sh.cell(row=max_row + record_line, column=update_index, value=sh.title)
                        break
                    # 从第2行开始取值
                    update_value = self.get_cell_value(row=row, col=pattern_indexs[i], sh=sh)
                    if sheet_info == 'titan' and new_update_texts and i in [_ for _ in range(refer_index)]:
                        update_value = new_update_texts[i]
                    if sheet_info == 'antivirus':
                        if i == refer_index:
                            sh_cell_text = list(
                                filter(lambda ip: False if (ip in invalid_infos or not ip.startswith('10.')) else ip,
                                       update_value.split()))
                            sh_cell_text.sort(key=lambda ip: ipaddress.ip_address(ip))
                            update_value = sh_cell_text[0]
                            try:
                                update_sh.cell(row=max_row + record_line, column=update_indexs[i + 1],
                                               value=sh_cell_text[1].strip())
                            except:
                                pass
                    # print(f"update_value1: {update_value}")
                    if sheet_info == 'host' and update_value:  # 对主机表的操作系统数据拼接
                        if i == refer_index:
                            update_value = f"{self.get_cell_value(row=row, col=pattern_indexs[i - 1], sh=sh)} {update_value}"
                        if i == 4:  # 对主机表的操作系统类型判断处理
                            for os_type in os_types:
                                if os_type.lower() in update_value.lower():
                                    update_value = os_type
                    if sheet_info == 'automation' and i == refer_index and update_value:  # 对自动化表的主机名数据拆分
                        try:
                            dn_ip = f'{update_value.split("_")[1]}'.strip()  # 先拿到带内IP
                            update_sh.cell(row=max_row + record_line, column=update_indexs[i - 1], value=dn_ip)
                        except:
                            pass
                        update_value = f'{update_value.split("_")[0]}'

                    # print(f"new_update_value: {update_value}")
                    if pattern_indexs.count(pattern_indexs[i]) > 1:  # 剔除重复比对数据不匹配情况
                        if i > pattern_indexs.index(pattern_indexs[i]):
                            continue

                    if type(update_value) is str:
                        update_value = update_value.strip()
                    update_sh.cell(row=max_row + record_line, column=update_indexs[i], value=update_value)

                record_line += 1

        print("<<<表{!r}更新完毕".format(update_sh.title))
        log.info("<<<表{!r}更新完毕".format(update_sh.title))

    def check_file_exist(self, file_path):
        '''查看文件路径是否存在'''
        if os.path.exists(file_path):
            print(f"找到需要解析的文件：{file_path}")
        else:
            print(f"未找到需要解析的文件：{file_path}")
            sys.exit(1)
        return file_path

    def get_index_list(self, pattern, sh) -> list:
        '''检查字段索引'''
        _indexs = []
        head_row_values = self.get_row_values(0, sh)
        if pattern:
            for _ in pattern:
                if _ in head_row_values:
                    _indexs.append(head_row_values.index(_) + 1)
            return _indexs

    def get_col_values(self, _index, pattern, sh):
        '''获取某列数据的所有内容，返回内容列表；需要留意的时，因为返回值被转为元组，改变了原索引的位置，故左移一位'''
        list_texts = []
        # 获取某列的全部数据列表
        for os_cell_text in tuple(sh.columns)[_index - 1]:
            # 剔除首行
            if pattern == os_cell_text.value:
                continue
            list_texts.append(os_cell_text.value)
        return list_texts

    def get_row_values(self, _index, sh):
        '''获取某行数据的所有内容，返回内容列表'''
        list_texts = []
        # 获取某列的全部数据列表
        for os_cell_text in tuple(sh.rows)[_index]:
            list_texts.append(os_cell_text.value)
        return list_texts

    def get_cell_value(self, row, col, sh):
        '''获取单元格的内容'''
        return sh.cell(row=row, column=col).value

    def clean_col_values(self, indexs, pattern, sh):
        # print([i.value for i in sh[1]])
        print("  开始清空{!r}已存在数据列{}--列号{}所有内容".format(sh.title, pattern, indexs))
        if indexs:
            if type(indexs) is list:
                print(f"  已清空列数据{pattern}--列号{indexs}")
                for i, index in enumerate(indexs):
                    for os_cell_text in tuple(sh.columns)[index - 1]:
                        # 剔除首行
                        if pattern[i] == os_cell_text.value:
                            continue
                        os_cell_text.value = None  # 只清空，不删除，否则会引起后列位移导致数据覆盖
            else:
                print(f"  已清空单列数据:{indexs}--列号:{pattern}")
                for os_cell_text in tuple(sh.columns)[indexs - 1]:
                    if pattern == os_cell_text.value:
                        continue
                    os_cell_text.value = None  # 只清空，不删除，否则会引起后列位移导致数据覆盖
                # sh.delete_cols(indexs)
                # sh.delete_cols(indexs)

            # 清理完加上头行数据标题
            self.set_row_head_value(indexs, pattern, sh)
            # print(f"头行数据1：{self.get_row_values(0, sh)}")

    def set_row_head_value(self, indexs, pattern, sh):
        # 增加头行信息
        # print(f"indexs:{indexs} <--> pattern:{pattern}")
        if indexs:
            if type(indexs) is list:
                print(f"  新增头行列表内所有列数据:{pattern}--列号:{indexs}")
                for i, info in enumerate(pattern):
                    sh.cell(row=1, column=indexs[i], value=info)
            else:
                print(f"  新增头行单列内数据:{pattern}--列号:{indexs}")
                sh.cell(row=1, column=indexs, value=pattern)

    def preset_head_value(self, pattern, sh, max_column):
        os_row_head = self.get_row_values(0, sh)
        i = 1
        for update_info in pattern:
            if update_info not in os_row_head:
                sh.cell(row=1, column=max_column + i, value=update_info)
                i += 1
        # print(self.get_row_values(0, sh))

    def recombine_texts(self, texts, pattern, sh):
        '''用于对ip组进行拆分重组'''
        new_texts = {}
        for hang, _ in enumerate(texts):
            if _ is None:
                print("  表{!r}的第{}行{!r}数据为{}".format(sh.title, hang + 2, pattern[0], _))
                continue
            new_keys = list(filter(lambda x: x != "", _.split(" ")))
            for key in new_keys:
                new_texts.update({key: hang})  # 忽略首行
        return new_texts

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(self.file_path)

    def __enter__(self):
        return self


class Excel_Host(BaseExcel):
    ...


if __name__ == '__main__':
    obj = Excel_Host()
    print(obj.yml_data)
    print(obj.config)
    print(obj.wb)
