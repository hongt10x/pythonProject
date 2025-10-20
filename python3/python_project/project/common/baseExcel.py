# -*- coding: utf-8 -*-
'''
@Software: PyCharm
@Project : python39
@File    : baseExcel.py
@Time    : 2024/5/30 10:40
@Author  : Echo Wang
'''

from openpyxl import load_workbook
import sys
import os
import time
import inspect
import ipaddress
import copy
from common.basePath import config_path, data_path
from utils.handle_yml import get_yml_data
from utils.handle_excel import get_excel_data
from utils.handle_loguru import log


def show_time(func):
    '''时间装饰器函数'''

    def inner(*args, **kwargs):
        start_time = time.monotonic()
        # log.info(f"\n主体函数执行开始的时间: %.2f秒秒" % start_time)
        ret = func(*args, **kwargs)
        end_time = time.monotonic()
        # log.info(f"主体函数执行结束的时间: %.2f秒秒" % end_time)
        print(f"  -->函数{func.__name__}调用后，执行总耗时：%.2f秒\n" % (end_time - start_time))
        log.info(f"  -->函数{func.__name__}调用后，执行总耗时：%.2f秒\n\n" % (end_time - start_time))
        return ret

    return inner


class BaseExcel(object):
    def __init__(self, config_yml='configDatas.yml'):
        # 确认配置文件有效
        config_file = self.check_file_exist(os.path.join(config_path, config_yml))
        # 获取配置文件内容
        self.yml_data = get_yml_data(config_file)
        # 获取excel路径
        self.file_path = self.check_file_exist(os.path.join(data_path, self.yml_data['filename']))
        # 读取表格数据
        self.wb = get_excel_data(self.file_path)
        # 定位哪个类调用基类，获取类名
        self.update_sheet = self.__class__.__name__.split("_")[-1].lower()
        # print(self.update_sheet)
        self.config = self.yml_data[f"update_{self.update_sheet}"]

    def compare_status(self, update_sh):
        ...
        compare_type = inspect.stack()[0][3]
        sheet_info = inspect.stack()[1][3].split("_")[-1]
        # print(f"sheet_info: {self.configs[sheet_info]['sheetName']}")
        # print(f"compare_type: {compare_type}")
        sh = self.wb[self.config[sheet_info]['sheetName']]
        # sh_max_row = sh.max_row + 1
        pattern = self.config[sheet_info][compare_type]['pattern']
        compare_logic = self.config[sheet_info][compare_type].get('compare_logic')
        # print(f"configs: {self.configs[sheet_info][compare_type]}")
        compare_column = self.config[sheet_info]['compare_column']
        update_info = self.config[sheet_info][compare_type]['update_info']
        invalid_infos = [None, "#N/A", "无", "#", "-", "    "]  # 单元格内容无效字符匹配
        # print(f"update_info: {update_info}")
        print(">>>开始将{!r}中的状态信息更新到{!r}中".format(sh.title, update_sh.title))
        log.info(">>>开始将{!r}中的状态信息更新到{!r}中".format(sh.title, update_sh.title))
        # 获取索引
        pattern_indexs = self.get_index_list(pattern, sh)
        update_indexs = self.get_index_list(update_info, update_sh)

        print(f"  查询{sh.title}表中列项数据:{pattern}<-->列项索引:{pattern_indexs}")
        print(f"  更新{update_sh.title}表中列项数据:{update_info}<-->列项索引:{update_indexs}")
        self.clean_col_values(update_indexs[compare_column:], update_info[compare_column:], update_sh)
        and_index_list = []
        update_texts = []
        update_max_len = 0
        # 因为先做了OS校对，可能导致实际比对行数小于sheet的最大行数，故不采用最大行max_row，取第一列比对行数
        sh_max_row = len([_ for _ in self.get_col_values(pattern_indexs[0], pattern[0], sh) if _ is not None])
        for col in range(compare_column):
            update_texts.extend(self.get_col_values(update_indexs[col], update_info[col], update_sh))
            if col == 0:
                update_max_len = len(update_texts)
        for i in range(2, sh_max_row + 2):
            # print("  与{!r}数据比对，当前执行第{}行...".format(self.configs[sheet_info]['sheetName'], i))
            and_index_list.clear()
            for column in range(compare_column):
                cell_value = self.get_cell_value(row=i, col=pattern_indexs[column], sh=sh)
                # print(f"cell_value1-->{i}:{cell_value}")
                # log.info(f"cell_value1-->{i}:{cell_value}")
                if cell_value not in invalid_infos and cell_value in update_texts:
                    v_index = update_texts.index(cell_value)
                    # log.info(f"行号:{i}的值:{cell_value} <--> 最大行:{update_max_len} - 原始索引:{v_index} ")
                    # 同一个列表合并数据时，需要考虑多列数据合并后的行号与索引无法对应问题，减去单列的列倍数还原行的索引
                    v_index = v_index if v_index < update_max_len else v_index - update_max_len * int(
                        v_index / update_max_len)
                    if compare_logic and compare_logic == 'AND':
                        and_index_list.append(v_index)
                        # print(f"and_index_list{column}: {and_index_list}")
                        continue
                    elif sheet_info in ['host', 'bastion']:  # 只标记天工或堡垒机是否纳管
                        update_sh.cell(row=v_index + 2, column=update_indexs[-1], value="已纳管")
                        break
                    else:
                        for k, idx in enumerate(update_indexs[compare_column:]):
                            cell_value = self.get_cell_value(row=i, col=pattern_indexs[compare_column + k], sh=sh)
                            # print(f"cell_value2-->{i}:{cell_value}")
                            # log.info(f"cell_value2-->{i}:{cell_value}")
                            update_sh.cell(row=v_index + 2, column=idx, value=cell_value)
                        break
            # print(f"and_index_list{i}: {and_index_list}")
            if and_index_list and len(and_index_list) == compare_column:
                for k, idx in enumerate(update_indexs[compare_column:]):  # 暂时保留该段代码
                    cell_value = self.get_cell_value(row=i, col=pattern_indexs[compare_column + k], sh=sh)
                    # print(f"cell_value3-->{i}:{cell_value}")
                    # log.info(f"cell_value3-->{i}:{cell_value}")
                    update_sh.cell(row=and_index_list[0] + 2, column=idx, value=cell_value)

        print("<<<表{!r}状态信息更新完毕".format(update_sh.title))
        log.info("<<<表{!r}状态信息更新完毕".format(update_sh.title))

    def compare_datas_other(self, update_sh):
        compare_type = inspect.stack()[0][3]
        sheet_info = inspect.stack()[1][3].split("_")[-1]
        # print(f"sheet_info:{sheet_info}")
        sh = self.wb[self.config[sheet_info]['sheetName']]
        # sh_max_row = sh.max_row
        # refer_index = self.configs[sheet_info][compare_type].get('refer_index')
        compare_column = self.config[sheet_info]['compare_column']
        pattern = self.config[sheet_info][compare_type]['pattern']
        update_info = self.config[sheet_info][compare_type]['update_info']
        print(">>>开始检查{!r}在{!r}中缺失的数据".format(sh.title, update_sh.title))
        log.info(">>>开始检查{!r}在{!r}中缺失的数据".format(sh.title, update_sh.title))
        pattern_indexs = self.get_index_list(pattern, sh)  # 获取关键列索引
        update_indexs = self.get_index_list(update_info, update_sh)  # 获取关键列索引
        print(f"  查询{sh.title}表中列项数据:{pattern}<-->列项索引:{pattern_indexs}")
        print(f"  更新{update_sh.title}表中列项数据:{update_info}<-->列项索引:{update_indexs}")
        invalid_infos = [None, "#N/A", "无", "#", "-", "    "]  # 单元格内容无效字符匹配
        missed_ips = {}  # 缺失ip列表
        cell_values = []
        update_texts = []
        for col in range(compare_column):
            update_texts.extend(self.get_col_values(update_indexs[col], update_info[col], update_sh))
        for row in range(2, sh.max_row + 1):
            # print("  与{!r}数据比对，当前执行第{}行...".format(self.configs[sheet_info]['sheetName'], row))
            flag = False
            cell_values.clear()
            for idx in pattern_indexs:
                cell_value = self.get_cell_value(row=row, col=idx, sh=sh)
                # 处理不存在或未匹配到的数据单元
                if cell_value not in invalid_infos and cell_value in update_texts:
                    flag = True
                    break
                else:
                    cell_values.append(cell_value)
            # missed_ips.update({row: tuple(cell_values)})
            if cell_values:  # 剔除已经包含的空的列表元素
                missed_ips.update({row: [_ for _ in cell_values]})
            if flag and missed_ips.get(row):
                # 找到数据后，如果记录长度小于循环次数相同，则删除missed_ips列表中的数据
                del missed_ips[row]
        # print(f"<<<检查完成，返回缺失的数据的行号和ip信息:\n  {missed_ips} \n")
        # log.info(f"<<<检查完成，返回缺失的数据的行号和ip信息:\n  {missed_ips}")
        # return missed_ips
        print(">>>开始将缺失数据，更新入{!r}的列{}中".format(update_sh.title, update_info))
        print(f"update_indexs:{update_indexs}")
        self.clean_col_values(update_indexs[compare_column:], update_info[compare_column:], update_sh)

        j = 1
        for num, ip in missed_ips.items():
            j += 1
            # print(num, ip,type(ip))
            update_sh.cell(row=j, column=update_indexs[compare_column:][0], value=num)
            update_sh.cell(row=j, column=update_indexs[compare_column:][1], value=ip[0])

        print("<<<表{!r}数据更新完毕".format(update_sh.title))
        log.info("<<<表{!r}数据更新完毕".format(update_sh.title))

    def compare_datas(self, update_sh, max_row):
        # 获取被调用函数名
        # print(inspect.stack()[0][3])
        # print(inspect.stack()[1][3])
        # 获取当前函数名
        compare_type = inspect.stack()[0][3]
        sheet_info = inspect.stack()[1][3].split("_")[-1]
        sh = self.wb[self.config[sheet_info]['sheetName']]
        print(sh)
        # sh_max_row = sh.max_row
        compare_column = self.config[sheet_info]['compare_column']
        pattern = self.config[sheet_info][compare_type]['pattern']
        update_info = self.config[sheet_info][compare_type]['update_info']
        refer_index = self.config[sheet_info][compare_type].get('refer_index')
        print(">>>开始准备将{!r}中独有的数据更新到{!r}中".format(sh.title, update_sh.title))
        log.info(">>>开始准备将{!r}中独有的数据更新到{!r}中".format(sh.title, update_sh.title))
        # 获取索引
        pattern_indexs = self.get_index_list(pattern, sh)
        update_indexs = self.get_index_list(update_info, update_sh)
        print(f"  查询{sh.title}表中列项数据:{pattern}<-->列项索引:{pattern_indexs}")
        print(f"  更新{update_sh.title}表中列项数据:{update_info}<-->列项索引:{update_indexs}")
        invalid_infos = [None, "#N/A", "无", "#", "-", "    "]  # 单元格内容无效字符匹配
        os_types = {'Linux': ['linux', 'kylin', 'red hat', 'centos'], 'Windows': ['windows'], 'AIX': ['aix'],
                    'HPUX': ['hp ux']}
        # 将防病毒库中独有的数据更新到OS表中
        # 因为先做了OS校对，可能导致实际比对行数小于sheet的最大行数，故不采用最大行max_row，取第一列比对行数
        sh_max_row = len([_ for _ in self.get_col_values(pattern_indexs[0], pattern[0], sh) if _ is not None])
        update_texts = []
        record_line = 1
        for col in range(compare_column):
            update_texts.extend(self.get_col_values(update_indexs[col], update_info[col], update_sh))
            print("update_texts：", update_texts)
        for row in range(2, sh_max_row + 2):
            # print("  与{!r}数据比对，当前执行第{}行...".format(self.configs[sheet_info]['sheetName'], row))
            match_times = 0
            for column in range(compare_column):
                ...
                cell_text = self.get_cell_value(row, pattern_indexs[column], sh)
                if cell_text not in invalid_infos and cell_text in update_texts:
                    break
                else:
                    match_times += 1  # 记录查询列数
            print("match_times: ", match_times)
            if match_times == compare_column:
                # 未匹配到任何列数据
                for i, update_index in enumerate(update_indexs):
                    if update_index == update_indexs[-1]:
                        # 最后一列更新数据注明数据来源，行往后追加
                        update_sh.cell(row=max_row + record_line, column=update_index, value=sh.title)
                        break
                    # 从第2行开始取值
                    update_value = self.get_cell_value(row=row, col=pattern_indexs[i], sh=sh)
                    if update_value and refer_index and refer_index == i:
                        if sheet_info == 'host':  # 给主机增加操作系统版本组合
                            try:
                                float(update_value)
                                next_update_value = self.get_cell_value(row=row, col=pattern_indexs[i + 1], sh=sh)
                                if next_update_value:
                                    update_value = f"{next_update_value} {update_value}"
                            except:
                                ...
                        elif sheet_info in ['titan', 'bastion']:  # 给青藤云、堡垒机增加操作系统类型匹配
                            for key, values in os_types.items():
                                for value in values:
                                    if value in update_value.lower():
                                        update_value = key
                    # print(f"update_value: {update_value}")
                    if type(update_value) is str:
                        update_value = update_value.strip()
                    if update_value:
                        update_sh.cell(row=max_row + record_line, column=update_indexs[i], value=update_value)
                record_line += 1
        print(">>>已完成将{!r}中独有的数据更新到{!r}中".format(sh.title, update_sh.title))
        log.info(">>>已完成将{!r}中独有的数据更新到{!r}中".format(sh.title, update_sh.title))

    def check_file_exist(self, file_path):
        '''查看文件路径是否存在'''
        if os.path.exists(file_path):
            print(f"找到需要解析的文件：{file_path}")
        else:
            print(f"未找到需要解析的文件：{file_path}")
            sys.exit(1)
        return file_path

    def get_index_list(self, pattern, sh) -> list:
        '''检查字段索引'''
        _indexs = []
        head_row_values = self.get_row_values(0, sh)
        if pattern:
            for _ in pattern:
                if _ in head_row_values:
                    _indexs.append(head_row_values.index(_) + 1)
            return _indexs

    def get_col_values(self, _index, pattern, sh):
        '''获取某列数据的所有内容，返回内容列表；需要留意的时，因为返回值被转为元组，改变了原索引的位置，故左移一位'''
        list_texts = []
        # 获取某列的全部数据列表
        for os_cell_text in tuple(sh.columns)[_index - 1]:
            # 剔除首行
            if pattern == os_cell_text.value:
                continue
            list_texts.append(os_cell_text.value)
        return list_texts

    def get_row_values(self, _index, sh):
        '''获取某行数据的所有内容，返回内容列表'''
        list_texts = []
        # 获取某行的全部数据列表
        for os_cell_text in tuple(sh.rows)[_index]:
            list_texts.append(os_cell_text.value)
        return list_texts

    def get_cell_value(self, row, col, sh):
        '''获取单元格的内容'''
        return sh.cell(row=row, column=col).value

    def clean_col_values(self, indexs, pattern, sh):
        # print([i.value for i in sh[1]])
        print("  开始清空{!r}已存在数据列{}--列号{}所有内容".format(sh.title, pattern, indexs))
        if indexs:
            if type(indexs) is list:
                print(f"  已清空列数据{pattern}--列号{indexs}")
                for i, index in enumerate(indexs):
                    for os_cell_text in tuple(sh.columns)[index - 1]:
                        # 剔除首行
                        if pattern[i] == os_cell_text.value:
                            continue
                        os_cell_text.value = None  # 只清空，不删除，否则会引起后列位移导致数据覆盖
            else:
                print(f"  已清空单列数据:{indexs}--列号:{pattern}")
                for os_cell_text in tuple(sh.columns)[indexs - 1]:
                    if pattern == os_cell_text.value:
                        continue
                    os_cell_text.value = None  # 只清空，不删除，否则会引起后列位移导致数据覆盖
                # sh.delete_cols(indexs)
                # sh.delete_cols(indexs)

            # 清理完加上头行数据标题
            self.set_row_head_value(indexs, pattern, sh)
            # print(f"头行数据1：{self.get_row_values(0, sh)}")

    def set_row_head_value(self, indexs, pattern, sh):
        # 增加头行信息
        # print(f"indexs:{indexs} <--> pattern:{pattern}")
        if indexs:
            if type(indexs) is list:
                print(f"  新增头行列表内所有列数据:{pattern}--列号:{indexs}")
                for i, info in enumerate(pattern):
                    sh.cell(row=1, column=indexs[i], value=info)
            else:
                print(f"  新增头行单列内数据:{pattern}--列号:{indexs}")
                sh.cell(row=1, column=indexs, value=pattern)

    def preset_head_value(self, pattern, sh, max_column):
        os_row_head = self.get_row_values(0, sh)
        i = 1
        for update_info in pattern:
            if update_info not in os_row_head:
                sh.cell(row=1, column=max_column + i, value=update_info)
                i += 1
        # print(self.get_row_values(0, sh))

    def recombine_texts(self, texts, pattern, sh):
        '''用于对ip组进行拆分重组'''
        new_texts = {}
        for hang, _ in enumerate(texts):
            if _ is None:
                print("  表{!r}的第{}行{!r}数据为{}".format(sh.title, hang + 2, pattern[0], _))
                continue
            new_keys = list(filter(lambda x: x != "", _.split(" ")))
            for key in new_keys:
                new_texts.update({key: hang})  # 忽略首行
        return new_texts

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(self.file_path)

    def __enter__(self):
        return self


class Excel_Host(BaseExcel):
    ...


if __name__ == '__main__':
    obj = Excel_Host()
    print(obj.yml_data)
    print(obj.config)
    print(obj.wb)
